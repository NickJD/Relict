"""Most Wanted List (MWL) taxonomy matching helpers.

The MWL workbook is treated as a taxonomy/priority overlay on top of the
standard PhyloSelect sequence assessment.  GTDB taxonomy remains the authoritative
input for MWL matching; other reference databases are still reported as
cross-checks through the existing multi-db classification columns.
"""

from __future__ import annotations

import csv
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional
from xml.etree import ElementTree as ET

from phyloselect.taxonomy import normalize_taxon_name, parse_taxon_string


RANK_ORDER = ['d', 'p', 'c', 'o', 'f', 'g', 's']
RANK_NAMES = {
    'd': 'domain',
    'k': 'kingdom',
    'p': 'phylum',
    'c': 'class',
    'o': 'order',
    'f': 'family',
    'g': 'genus',
    's': 'species',
}
RANK_SCORES = {
    'd': 10.0,
    'k': 10.0,
    'p': 30.0,
    'c': 45.0,
    'o': 60.0,
    'f': 75.0,
    'g': 90.0,
    's': 100.0,
}


@dataclass
class MWLEntry:
    mwl_id: str
    hierarchy: str
    role: str = ''
    ranks: Dict[str, set[str]] = field(default_factory=dict)
    unranked_terms: set[str] = field(default_factory=set)


@dataclass
class MWLMatch:
    entry: MWLEntry
    rank: str
    taxon: str
    taxonomic_score: float
    identity: Optional[float]
    mwl_score: float


def _safe_text(value: object) -> str:
    if value is None:
        return ''
    return str(value).replace('\t', ' ').replace('\n', ' ').replace('\r', ' ').strip()


def _rank_key(rank: str) -> str:
    rank = str(rank or '').strip().lower()
    aliases = {
        'domain': 'd',
        'kingdom': 'k',
        'phylum': 'p',
        'class': 'c',
        'order': 'o',
        'family': 'f',
        'genus': 'g',
        'species': 's',
    }
    return aliases.get(rank, rank[:1] if rank else '')


def _rank_allowed(rank: str, min_rank: str, entry: MWLEntry) -> bool:
    rank = _rank_key(rank)
    min_rank = _rank_key(min_rank or 'p')
    if rank in ('d', 'k') and _entry_has_only_domain(entry):
        return True
    try:
        return RANK_ORDER.index(rank if rank != 'k' else 'd') >= RANK_ORDER.index(min_rank if min_rank != 'k' else 'd')
    except ValueError:
        return False


def _entry_has_only_domain(entry: MWLEntry) -> bool:
    ranked = {r for r, vals in entry.ranks.items() if vals}
    return ranked.issubset({'d', 'k'})


def _add_rank_value(ranks: Dict[str, set[str]], rank: str, value: str) -> None:
    key = _rank_key(rank)
    norm = normalize_taxon_name(value)
    if key and norm:
        ranks.setdefault(key, set()).add(norm)


def _parse_entry_terms(hierarchy: str) -> tuple[Dict[str, set[str]], set[str]]:
    """Extract ranked and unranked MWL match terms from a hierarchy string."""
    ranks: Dict[str, set[str]] = {}
    unranked: set[str] = set()
    hierarchy = _safe_text(hierarchy)
    if not hierarchy:
        return ranks, unranked

    # Main lineage outside parenthetical "including" clauses.
    main_lineage = re.sub(r'\([^)]*\)', '', hierarchy)
    for part in main_lineage.split(';'):
        part = part.strip()
        if '__' not in part:
            continue
        rank, value = part.split('__', 1)
        _add_rank_value(ranks, rank, value)

    # Parenthetical clauses often contain additional genera/species.
    for inner in re.findall(r'\(([^)]*)\)', hierarchy):
        inner = re.sub(r'^\s*including\s*:?\s*', '', inner, flags=re.IGNORECASE)
        for match in re.finditer(r'([dkpcofgs])__([^;,()]+)', inner):
            _add_rank_value(ranks, match.group(1), match.group(2))

        cleaned = re.sub(r'([dkpcofgs])__[^;,()]+', ' ', inner)
        for token in re.split(r'[;,]', cleaned):
            token = token.strip().strip(':')
            if not token or token.lower() == 'including':
                continue
            norm = normalize_taxon_name(token)
            if norm:
                unranked.add(norm)

    return ranks, unranked


def _rows_from_tsv_or_csv(path: str) -> List[List[object]]:
    delimiter = '\t' if str(path).lower().endswith(('.tsv', '.txt')) else ','
    with open(path, newline='') as fh:
        return [row for row in csv.reader(fh, delimiter=delimiter)]


def _xlsx_rows_openpyxl(path: str, sheet_name: str) -> Optional[List[List[object]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"MWL workbook does not contain sheet '{sheet_name}'. Available: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]


def _xlsx_rows_stdlib(path: str, sheet_name: str) -> List[List[object]]:
    """Read simple cell values from an .xlsx workbook using only stdlib."""
    ns = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'rel': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'pkgrel': 'http://schemas.openxmlformats.org/package/2006/relationships',
    }
    with zipfile.ZipFile(path) as zf:
        shared: List[str] = []
        if 'xl/sharedStrings.xml' in zf.namelist():
            root = ET.fromstring(zf.read('xl/sharedStrings.xml'))
            for si in root.findall('main:si', ns):
                texts = [t.text or '' for t in si.findall('.//main:t', ns)]
                shared.append(''.join(texts))

        wb_root = ET.fromstring(zf.read('xl/workbook.xml'))
        rel_root = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        rels = {
            rel.attrib['Id']: rel.attrib['Target']
            for rel in rel_root.findall('pkgrel:Relationship', ns)
        }

        target = None
        for sheet in wb_root.findall('.//main:sheet', ns):
            if sheet.attrib.get('name') == sheet_name:
                target = rels.get(sheet.attrib.get(f"{{{ns['rel']}}}id"))
                break
        if target is None:
            names = [s.attrib.get('name', '') for s in wb_root.findall('.//main:sheet', ns)]
            raise ValueError(f"MWL workbook does not contain sheet '{sheet_name}'. Available: {', '.join(names)}")
        if target.startswith('/'):
            target = target.lstrip('/')
        elif not target.startswith('xl/'):
            target = 'xl/' + target.lstrip('/')

        sheet_root = ET.fromstring(zf.read(target))
        rows: List[List[object]] = []
        for row in sheet_root.findall('.//main:row', ns):
            values: Dict[int, object] = {}
            for cell in row.findall('main:c', ns):
                ref = cell.attrib.get('r', '')
                col_letters = ''.join(ch for ch in ref if ch.isalpha())
                col_idx = _column_index(col_letters)
                ctype = cell.attrib.get('t')
                value = ''
                if ctype == 'inlineStr':
                    value = ''.join(t.text or '' for t in cell.findall('.//main:t', ns))
                else:
                    v = cell.find('main:v', ns)
                    if v is not None and v.text is not None:
                        value = shared[int(v.text)] if ctype == 's' else v.text
                values[col_idx] = value
            if values:
                rows.append([values.get(i, '') for i in range(1, max(values) + 1)])
        return rows


def _column_index(letters: str) -> int:
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx or 1


def _load_rows(path: str, sheet_name: str) -> List[List[object]]:
    lower = str(path).lower()
    if lower.endswith('.xlsx'):
        rows = _xlsx_rows_openpyxl(path, sheet_name)
        return rows if rows is not None else _xlsx_rows_stdlib(path, sheet_name)
    return _rows_from_tsv_or_csv(path)


def load_mwl_entries(path: str, sheet_name: str = 'MWL_V1') -> List[MWLEntry]:
    rows = _load_rows(path, sheet_name)
    if not rows:
        return []
    header = [_safe_text(v).lower() for v in rows[0]]

    def _find_col(candidates: Iterable[str], default: int) -> int:
        for cand in candidates:
            for idx, label in enumerate(header):
                if cand in label:
                    return idx
        return default

    id_idx = _find_col(['mwl', 'id'], 0)
    hierarchy_idx = _find_col(['hierarchy', 'taxa', 'taxonomic'], 1)
    role_idx = _find_col(['functional', 'metabolic', 'role'], 2)

    entries: List[MWLEntry] = []
    for row in rows[1:]:
        if not row:
            continue
        mwl_id = _safe_text(row[id_idx] if id_idx < len(row) else '')
        hierarchy = _safe_text(row[hierarchy_idx] if hierarchy_idx < len(row) else '')
        role = _safe_text(row[role_idx] if role_idx < len(row) else '')
        if not mwl_id or not hierarchy:
            continue
        ranks, unranked = _parse_entry_terms(hierarchy)
        entries.append(MWLEntry(mwl_id=mwl_id, hierarchy=hierarchy, role=role, ranks=ranks, unranked_terms=unranked))
    return entries


def match_taxonomy_to_mwl(
    taxonomy: str,
    entries: List[MWLEntry],
    *,
    identity: Optional[float] = None,
    min_rank: str = 'p',
) -> Optional[MWLMatch]:
    parsed = parse_taxon_string(taxonomy)
    if not parsed:
        return None

    best: Optional[MWLMatch] = None
    for entry in entries:
        for rank in RANK_ORDER:
            value = parsed.get(rank)
            if not value:
                continue
            norm = normalize_taxon_name(value)
            rank_hit = norm in entry.ranks.get(rank, set()) or norm in entry.unranked_terms
            if not rank_hit or not _rank_allowed(rank, min_rank, entry):
                continue

            tax_score = RANK_SCORES.get(rank, 0.0)
            ident_factor = max(0.0, min(float(identity), 100.0)) / 100.0 if identity is not None else 1.0
            mwl_score = round(tax_score * ident_factor, 2)
            candidate = MWLMatch(
                entry=entry,
                rank=rank,
                taxon=value,
                taxonomic_score=tax_score,
                identity=identity,
                mwl_score=mwl_score,
            )
            if best is None or (candidate.mwl_score, candidate.taxonomic_score, candidate.entry.mwl_id) > (best.mwl_score, best.taxonomic_score, best.entry.mwl_id):
                best = candidate
    return best


def _float_or_none(value: object) -> Optional[float]:
    try:
        text = str(value).strip()
        if not text or text.upper() == 'NA':
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def add_evaluation_scores(rows: List[dict]) -> None:
    for row in rows:
        mwl_score = _float_or_none(row.get('mwl_score')) or 0.0
        investigation = _float_or_none(row.get('investigation_score'))
        novelty = _float_or_none(row.get('novelty_score')) or 0.0
        if investigation is not None:
            score = (0.75 * investigation) + (0.25 * mwl_score)
        else:
            score = (0.65 * novelty) + (0.35 * mwl_score)
        row['evaluation_score'] = f"{round(min(100.0, score), 2):.2f}"


def annotate_assessment_rows(
    rows: List[dict],
    entries: List[MWLEntry],
    *,
    min_rank: str = 'p',
) -> List[dict]:
    for row in rows:
        identity = _float_or_none(row.get('classification_identity'))
        match = match_taxonomy_to_mwl(str(row.get('taxonomy') or ''), entries, identity=identity, min_rank=min_rank)
        if match is None:
            row.update({
                'mwl_match': 'No',
                'mwl_id': 'NA',
                'mwl_matched_rank': 'NA',
                'mwl_matched_taxon': 'NA',
                'mwl_taxonomic_score': '0.00',
                'mwl_identity': f"{identity:.2f}" if identity is not None else 'NA',
                'mwl_score': '0.00',
                'mwl_role': 'NA',
            })
            continue
        row.update({
            'mwl_match': 'Yes',
            'mwl_id': match.entry.mwl_id,
            'mwl_matched_rank': RANK_NAMES.get(match.rank, match.rank),
            'mwl_matched_taxon': _safe_text(match.taxon),
            'mwl_taxonomic_score': f"{match.taxonomic_score:.2f}",
            'mwl_identity': f"{match.identity:.2f}" if match.identity is not None else 'NA',
            'mwl_score': f"{match.mwl_score:.2f}",
            'mwl_role': _safe_text(match.entry.role),
        })
    add_evaluation_scores(rows)
    return rows


def write_mwl_matches_tsv(path: str | Path, rows: List[dict], *, matches_only: bool = True) -> str:
    p = Path(path)
    header = [
        'ID',
        'Taxonomy',
        'ClassificationIdentity',
        'MWLMatch',
        'MWLID',
        'MWLMatchedRank',
        'MWLMatchedTaxon',
        'MWLTaxonomicScore',
        'MWLIdentity',
        'MWLScore',
        'EvaluationScore',
        'MWLRole',
    ]
    with open(p, 'w', newline='') as fh:
        fh.write('\t'.join(header) + '\n')
        for row in rows:
            if matches_only and row.get('mwl_match') != 'Yes':
                continue
            fh.write('\t'.join([
                _safe_text(row.get('id')),
                _safe_text(row.get('taxonomy')),
                _safe_text(row.get('classification_identity')),
                _safe_text(row.get('mwl_match')),
                _safe_text(row.get('mwl_id')),
                _safe_text(row.get('mwl_matched_rank')),
                _safe_text(row.get('mwl_matched_taxon')),
                _safe_text(row.get('mwl_taxonomic_score')),
                _safe_text(row.get('mwl_identity')),
                _safe_text(row.get('mwl_score')),
                _safe_text(row.get('evaluation_score')),
                _safe_text(row.get('mwl_role')),
            ]) + '\n')
    return str(p)
